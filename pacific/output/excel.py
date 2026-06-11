"""PACIFIC — Excel/spreadsheet output generation."""

from datetime import datetime
from pathlib import Path
from typing import Optional

from rich.console import Console

from pacific.config import load_config

console = Console()


def export_to_excel(
    data: dict,
    filename: Optional[str] = None,
    sheet_name: str = "Pacific Analysis",
) -> str:
    """Export data to a formatted Excel spreadsheet.

    Args:
        data: Dict with 'columns' (list) and 'rows' (list of lists),
              or a pandas DataFrame.
        filename: Output filename. Auto-generated if None.
        sheet_name: Name of the Excel sheet.

    Returns:
        Path to saved Excel file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print("[red]Missing dependency: pip install openpyxl[/red]")
        return ""

    cfg = load_config()
    out_dir = Path(cfg.get("output_dir", "~/.pacific/outputs")).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pacific_report_{ts}.xlsx"

    filepath = out_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Handle dict or DataFrame
    try:
        import pandas as pd
        if isinstance(data, pd.DataFrame):
            columns = list(data.columns)
            rows = data.values.tolist()
        else:
            columns = data.get("columns", [])
            rows = data.get("rows", [])
    except ImportError:
        columns = data.get("columns", [])
        rows = data.get("rows", [])

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    title_cell = ws.cell(row=1, column=1, value=f"PACIFIC — {sheet_name}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="808080", size=9)

    # Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    for row_idx, row in enumerate(rows, 5):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")

    # Auto-width
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_width = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(4, len(rows) + 6)
        )
        ws.column_dimensions[col_letter].width = min(max_width + 4, 40)

    wb.save(filepath)
    console.print(f"[green]✓ Excel saved:[/green] {filepath}")
    return str(filepath)


def stock_data_to_excel(
    ticker: str,
    period: str = "3mo",
    filename: Optional[str] = None,
) -> str:
    """Download stock data and export to Excel."""
    try:
        import yfinance as yf
        import pandas as pd
    except ImportError:
        console.print("[red]Missing dependency: pip install yfinance openpyxl[/red]")
        return ""

    console.print(f"[dim]Fetching {ticker} data ({period})…[/dim]")
    data = yf.download(ticker, period=period, progress=False)
    if data.empty:
        console.print(f"[red]No data for {ticker}[/red]")
        return ""

    if hasattr(data.columns, 'levels') and len(data.columns.levels) > 1:
        data.columns = data.columns.get_level_values(0)

    data = data.reset_index()
    data["Date"] = data["Date"].dt.strftime("%Y-%m-%d")

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{ticker}_data_{ts}.xlsx"

    return export_to_excel(data, filename=filename, sheet_name=f"{ticker} Data")

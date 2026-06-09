"""
Pacific CLI — Export: PDF reports, Excel spreadsheets, JSON files.

All exports saved to ~/.pacific/outputs/ with PACIFIC branding.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from .config import load_config

# ANSI colors
RESET = "\033[0m"
GREEN = "\033[1;32m"
RED = "\033[1;31m"
DIM = "\033[2m"

# Unicode → ASCII-safe replacements for PDF core fonts
_UNICODE_SUBS = {
    "\u2022": "-", "\u2013": "-", "\u2014": "--",
    "\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"',
    "\u2026": "...", "\u2192": "->", "\u2190": "<-",
    "\u2264": "<=", "\u2265": ">=", "\u2260": "!=",
    "\u00d7": "x", "\u00f7": "/", "\u221a": "sqrt",
    "\u03b1": "alpha", "\u03b2": "beta", "\u03b3": "gamma",
    "\u03b4": "delta", "\u03c3": "sigma", "\u03bc": "mu",
    "\u03c0": "pi", "\u0394": "Delta", "\u03a3": "Sigma",
}


def _safe_text(text: str) -> str:
    """Replace unsupported Unicode for PDF core fonts."""
    for orig, repl in _UNICODE_SUBS.items():
        text = text.replace(orig, repl)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _get_output_dir() -> Path:
    cfg = load_config()
    out_dir = Path(cfg.get("output_dir", str(Path.home() / ".pacific" / "outputs")))
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


# ─── PDF Export ──────────────────────────────────────────────────────

def analysis_to_pdf(title: str, content: str, filename: Optional[str] = None) -> str:
    """Export analysis text to a styled PDF report."""
    try:
        from fpdf import FPDF
    except ImportError:
        print(f"{RED}Missing: pip install fpdf2{RESET}")
        return ""

    out_dir = _get_output_dir()
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = title.lower().replace(" ", "_")[:30]
        filename = f"pacific_{safe_title}_{ts}.pdf"

    filepath = out_dir / filename

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Header
    pdf.set_fill_color(27, 79, 114)
    pdf.rect(0, 0, 210, 30, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_xy(10, 8)
    pdf.cell(0, 10, "PACIFIC", ln=False)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_xy(10, 18)
    pdf.cell(0, 6, "Elite Financial Intelligence | Powered by GRRN")

    # Title
    pdf.set_text_color(27, 79, 114)
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_xy(10, 38)
    pdf.cell(0, 10, _safe_text(title), ln=True)

    # Timestamp
    pdf.set_text_color(128, 128, 128)
    pdf.set_font("Helvetica", "I", 9)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(6)

    # Content
    pdf.set_text_color(33, 33, 33)
    pdf.set_font("Helvetica", "", 10)

    safe_content = _safe_text(content)
    for line in safe_content.split("\n"):
        line = line.rstrip()
        if line.startswith("# "):
            pdf.set_font("Helvetica", "B", 13)
            pdf.cell(0, 8, line[2:], ln=True)
            pdf.set_font("Helvetica", "", 10)
        elif line.startswith("## "):
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 7, line[3:], ln=True)
            pdf.set_font("Helvetica", "", 10)
        elif line.startswith("**") and line.endswith("**"):
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, line.strip("*"), ln=True)
            pdf.set_font("Helvetica", "", 10)
        elif line.startswith("- ") or line.startswith("* "):
            pdf.cell(10, 6, "")
            pdf.cell(0, 6, f"  {line}", ln=True)
        else:
            pdf.multi_cell(0, 5, line)

    # Footer
    pdf.set_y(-20)
    pdf.set_text_color(128, 128, 128)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 6, "Pacific CLI | https://pacific.grrn.io", align="C")

    pdf.output(str(filepath))
    print(f"{GREEN}✓ PDF saved:{RESET} {filepath}")
    return str(filepath)


# ─── JSON Export ─────────────────────────────────────────────────────

def analysis_to_json(title: str, content: str, filename: Optional[str] = None) -> str:
    """Export analysis text to a JSON file with metadata."""
    out_dir = _get_output_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not filename:
        safe = title.lower().replace(" ", "_")[:30]
        filename = f"pacific_{safe}_{ts}.json"
    if not filename.endswith(".json"):
        filename += ".json"

    filepath = out_dir / filename

    envelope = {
        "meta": {
            "generator": "Pacific CLI — Elite Financial Intelligence",
            "version": "1.0.0",
            "generated_at": datetime.now().isoformat(),
            "title": title,
        },
        "data": {"analysis": content},
    }

    filepath.write_text(json.dumps(envelope, indent=2, default=str), encoding="utf-8")
    print(f"{GREEN}✓ JSON saved:{RESET} {filepath}")
    return str(filepath)


def stock_data_to_json(ticker: str, period: str = "3mo", filename: Optional[str] = None) -> str:
    """Download stock data and export to JSON."""
    try:
        import yfinance as yf
    except ImportError:
        print(f"{RED}Missing: pip install yfinance{RESET}")
        return ""

    print(f"{DIM}Fetching {ticker} data ({period})...{RESET}")
    df = yf.download(ticker.upper(), period=period, progress=False)
    if df.empty:
        print(f"{RED}No data for {ticker}{RESET}")
        return ""

    if hasattr(df.columns, "levels") and len(df.columns.levels) > 1:
        df.columns = df.columns.get_level_values(0)

    df = df.reset_index()
    df["Date"] = df["Date"].dt.strftime("%Y-%m-%d")

    out_dir = _get_output_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not filename:
        filename = f"{ticker.upper()}_data_{ts}.json"

    filepath = out_dir / filename

    envelope = {
        "meta": {
            "generator": "Pacific CLI",
            "ticker": ticker.upper(),
            "period": period,
            "generated_at": datetime.now().isoformat(),
        },
        "data": json.loads(df.to_json(orient="records", date_format="iso")),
    }

    filepath.write_text(json.dumps(envelope, indent=2, default=str), encoding="utf-8")
    print(f"{GREEN}✓ JSON saved:{RESET} {filepath}")
    return str(filepath)


# ─── Excel Export ────────────────────────────────────────────────────

def stock_data_to_excel(ticker: str, period: str = "3mo", filename: Optional[str] = None) -> str:
    """Download stock data and export to a formatted Excel spreadsheet."""
    try:
        import yfinance as yf
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"{RED}Missing: pip install yfinance openpyxl{RESET}")
        return ""

    print(f"{DIM}Fetching {ticker.upper()} data ({period})...{RESET}")
    df = yf.download(ticker.upper(), period=period, progress=False)
    if df.empty:
        print(f"{RED}No data for {ticker.upper()}{RESET}")
        return ""

    if hasattr(df.columns, "levels") and len(df.columns.levels) > 1:
        df.columns = df.columns.get_level_values(0)

    df = df.reset_index()

    out_dir = _get_output_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not filename:
        filename = f"{ticker.upper()}_data_{ts}.xlsx"

    filepath = out_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{ticker.upper()} Data"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    columns = list(df.columns)
    rows = df.values.tolist()

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    title_cell = ws.cell(row=1, column=1, value=f"PACIFIC — {ticker.upper()} Stock Data")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Period: {period}")
    ws.cell(row=2, column=1).font = Font(italic=True, color="808080", size=9)

    # Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=str(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    for row_idx, row in enumerate(rows, 5):
        for col_idx, value in enumerate(row, 1):
            # Convert numpy/pandas types
            if hasattr(value, "item"):
                value = value.item()
            elif hasattr(value, "strftime"):
                value = value.strftime("%Y-%m-%d")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"

    # Auto-width
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_width = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(4, min(len(rows) + 6, 105))
        )
        ws.column_dimensions[col_letter].width = min(max_width + 4, 40)

    wb.save(filepath)
    print(f"{GREEN}✓ Excel saved:{RESET} {filepath}")
    return str(filepath)
